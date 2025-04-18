import pandas as pd
import re
import openpyxl
from openpyxl import workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import borders
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import numpy as np

#-----------------------------------------------------------------------------
# Data Processing


df1 = pd.read_excel('Data/Raw_Data_Input.xlsx')
df1.rename(columns = {"Unnamed: 0" : " "}, inplace = True)
data_list_new = []
component_list = []

# Obtaining index numbers for Phase, T, P, Enthalpy, and Density
for i in df1.index:
    if df1[' '].iloc[i] == 'Phase':
        Phase_Index = i
        data_list_new.append(Phase_Index)
        break
for i in df1.index:
    if df1[' '].iloc[i] == 'Temperature':
        Temperature_Index = i
        data_list_new.append(Temperature_Index)
        break
for i in df1.index:
    if df1[' '].iloc[i] == 'Pressure':
        Pressure_Index = i
        data_list_new.append(Pressure_Index)
        break
for i in df1.index:
    if df1[' '].iloc[i] == 'Molar Enthalpy':
        Molar_Enthalpy_Index = i
        data_list_new.append(Molar_Enthalpy_Index)
        break
for i in df1.index:
    if df1[' '].iloc[i] == 'Mass Enthalpy':
        Mass_Enthalpy_Index = i
        data_list_new.append(Mass_Enthalpy_Index)
        break
for i in df1.index:
    if df1[' '].iloc[i] == 'Mass Density':
        Mass_Density_Index = i
        data_list_new.append(Mass_Density_Index)
        break

# Obtaining index number for total molar flow
for i in df1.index:
    if df1[' '].iloc[i] == 'Mole Flows':
        Molar_Flow_Index = i
        data_list_new.append(Molar_Flow_Index)
        break
    else:
        pass

#Obtain index numbers for mass flows
for i in df1.index:
    if df1[' '].iloc[i] == 'Mass Flows':
        Mass_Flow_Unit = df1['Units'].iloc[i]
for i in df1.index:
    if df1[' '].iloc[i] == 'Vapor Phase':
        break
    elif df1['Units'].iloc[i] == Mass_Flow_Unit:
        data_list_new.append(i)
        component_list.append(i)
    else:
        pass


df1 = df1[df1.index.isin(data_list_new)]                            #filtering out all the rows you want
#df1.columns = ["" + str(col) for col in df1.columns]           # placing a ' on stream names to stop dates showing up
#df1.rename(columns = {"Unnamed: 0" : " "}, inplace = True)     # Renaming unwanted column name

#Renaming parameters column
df1.loc[24, ' '] = 'Mole Flow'
df1.loc[38, ' '] = 'Mass Flow'
component_list.pop(0)
for i in component_list:
    df1.loc[i, ' '] = df1.loc[i, ' '] + ' Mass Flow'


for column in df1:
    if column == 'Units':
        pass
    else:
        df1.loc[7, column] = str(df1.loc[7, column])
        if df1.loc[7, column] == 'nan':
            df1.loc[7, column] = 'Vapour/Liquid Phase'
        else:
            pass

#converting numbers to strings
def convert_string(x):
    x = str(x)
    if re.search('\d', x):
        x = float(x)
        x = round(x, 1)
        return x
    else:
        return x

for item in df1:
    if item == " ":
        pass
    elif item == "Units":
        pass
    else:
        df1[item] = df1[item].apply(convert_string)

df1.set_index(' ', inplace = True)
df1 = df1.T

#column_to_move = df1.pop('Fluid Density')
#df1.insert(13, 'Fluid Density', column_to_move)

df1 = df1.style.set_properties(**{'text-align' : 'center'})
df1.to_excel('Outputs/Results.xlsx', index = True)


#----------------------------------------------------------------------------
# Formatting Excel Spreadsheet

wb = load_workbook('Outputs/Results.xlsx')
ws = wb.active

# Setting columns width properties
for col in ws.columns:
     max_length = 0
     column = col[0].column_letter # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 1) * 1.2
     ws.column_dimensions[column].width = adjusted_width

y = len(df1.index) + 4
x = len(df1.columns) + 1
z = get_column_letter(len(df1.columns) + 1)

ws.insert_rows(0)
ws.insert_rows(0)
ws.insert_rows(0)
ws['A1'] = 'Mass & Energy Balance'
ws.merge_cells('A1:' + str(z) + '1')

font_style_1 = Font(size = '24')
ws['A1'].alignment = Alignment(horizontal = "center")
ws['A1'].font = font_style_1

border1 = borders.Side(style = None, color = 'FF000000', border_style = 'thin')
border0 = borders.Side(style = None, color = None, border_style = None)
thin = Border(left = border1, right = border1, bottom = border1, top = border1)

for row in ws.iter_rows(min_row = 5, min_col = 2, max_row = y, max_col = x):
    for cell in row:
        cell.border = thin


print('Script Sucessfully Completed')
#time.sleep(3)
wb.save('Outputs/Results.xlsx')



# ME_Balance_Programme_Code.py
